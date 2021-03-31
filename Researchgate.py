from bs4 import BeautifulSoup as bs
from urllib.request import urlopen
from requests_html import HTMLSession
from selenium import webdriver
import requests
import pandas as pd
import xlrd
from openpyxl import *

wb1=load_workbook("C:\Dashboard\Full Time faculty details_Qualification and Classification.xlsx")
ws=wb1["Full-Time"]
loc = ("C:\Dashboard\Full Time faculty details_Qualification and Classification.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
for i in range(1,33):
    k=22
    A = sheet.cell_value(i, 14)
    wcell0=ws.cell(i+1,1)
    #B = sheet.cell_value(i,2)
    #wcell0.value = B
    if A == 0:
        k+=1
    if A != 0:
        page_response = requests.get(A)
        page_content = bs(page_response.content, "html.parser")
        detail = page_content.find("div",attrs={"class":"nova-c-card"})
        info = detail.find_all("div",attrs={"class":"nova-e-text--size-xl"})
        for j in info:
            print(j.text)
            k+=1
            wcell1 = ws.cell(i+1,k)
            wcell1.value = j.text
wb1.save("C:\Dashboard\Full Time faculty details_Qualification and Classification.xlsx")    
#print(reads)
